from __future__ import annotations
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    DATA_FILE_PATH, SHEET_TX, SHEET_WC, SHEET_LOC, SHEET_LISTS, SHEET_CURRENT
)
from db.models import Transaction


def ensure_workbook_exists(path: Path = DATA_FILE_PATH) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")


def _find_header_column(ws: Worksheet, header_text: str) -> int:
    # header is in row 1, exact match
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and v.strip() == header_text:
            return col
    raise KeyError(f"Missing header '{header_text}' in sheet '{ws.title}'")


def list_active_wheelchair_ids(path: Path = DATA_FILE_PATH) -> List[str]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_WC]

    out: List[str] = []
    # A = ID, I = Active (כן/לא). If empty => treat as active (per VBA)
    for r in range(2, ws.max_row + 1):
        wid = ws.cell(r, 1).value
        if wid is None:
            continue
        wid_s = str(wid).strip()
        if not wid_s:
            continue
        active = ws.cell(r, 9).value  # column I
        active_s = str(active).strip().lower() if active is not None else ""
        if active_s in ("", "כן", "true", "1", "y", "yes"):
            out.append(wid_s)
    wb.close()
    return out


def list_locations(path: Path = DATA_FILE_PATH) -> List[Dict[str, Any]]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_LOC]

    c_id = _find_header_column(ws, "LocationID")
    c_name = _find_header_column(ws, "LocationName")

    out: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        loc_id = ws.cell(r, c_id).value
        if loc_id is None or str(loc_id).strip() == "":
            continue
        loc_name = ws.cell(r, c_name).value
        out.append({"LocationID": int(loc_id), "LocationName": str(loc_name or "").strip()})
    wb.close()
    return out


def list_enums(path: Path = DATA_FILE_PATH) -> Dict[str, List[str]]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_LISTS]

    txn: List[str] = []
    status: List[str] = []
    cond: List[str] = []

    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if a is not None and str(a).strip():
            txn.append(str(a).strip())
        if b is not None and str(b).strip():
            status.append(str(b).strip())
        if c is not None and str(c).strip():
            cond.append(str(c).strip())

    wb.close()
    return {"TxnType": txn, "StatusAfter": status, "ConditionAfter": cond}


def get_current_location_id(wheelchair_id: str, path: Path = DATA_FILE_PATH) -> Optional[int]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_CURRENT]

    target = str(wheelchair_id).strip()
    if not target:
        wb.close()
        return None

    # find in col A
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        if str(v).strip() == target:
            cur = ws.cell(r, 7).value  # Column 7 = CurrentLocationID (per VBA)
            wb.close()
            return int(cur) if cur is not None and str(cur).strip() != "" else None

    wb.close()
    return None


def append_transaction(tx: Transaction, path: Path = DATA_FILE_PATH) -> int:
    ensure_workbook_exists(path)
    wb = load_workbook(path)
    ws = wb[SHEET_TX]

    # Next empty row in col A
    last = ws.max_row
    while last >= 2 and (ws.cell(last, 1).value is None or str(ws.cell(last, 1).value).strip() == ""):
        last -= 1
    new_row = 2 if last < 2 else last + 1

    # Next TxnID = last A + 1
    if last < 2 or ws.cell(last, 1).value in (None, ""):
        txn_id = 1
    else:
        txn_id = int(ws.cell(last, 1).value) + 1

    ws.cell(new_row, 1).value = txn_id
    ws.cell(new_row, 2).value = tx.timestamp
    ws.cell(new_row, 2).number_format = "dd/mm/yyyy hh:mm"
    ws.cell(new_row, 3).value = tx.created_by
    ws.cell(new_row, 4).value = tx.wheelchair_id
    ws.cell(new_row, 5).value = tx.txn_type
    ws.cell(new_row, 6).value = tx.from_location_id
    ws.cell(new_row, 7).value = tx.to_location_id
    ws.cell(new_row, 8).value = tx.status_after
    ws.cell(new_row, 9).value = tx.condition_after or ""
    ws.cell(new_row,10).value = tx.patient_name or ""
    ws.cell(new_row,11).value = tx.notes or ""

    # SortKey formula in col L
    ws.cell(new_row,12).value = f'=IF(AND(A{new_row}<>"",B{new_row}<>""),B{new_row}+A{new_row}/1000000,"")'

    wb.save(path)
    wb.close()
    return txn_id


def count_txn_type_in_range(txn_type: str, start: date, end: date, path: Path = DATA_FILE_PATH) -> int:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_TX]

    c_ts = _find_header_column(ws, "Timestamp")
    c_type = _find_header_column(ws, "TxnType")

    cnt = 0
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, c_type).value
        if str(t or "").strip() != txn_type:
            continue
        vts = ws.cell(r, c_ts).value
        if isinstance(vts, datetime):
            d = vts.date()
        else:
            try:
                d = datetime.fromisoformat(str(vts)).date()
            except Exception:
                continue
        if start <= d <= end:
            cnt += 1

    wb.close()
    return cnt


def status_aggregation_from_current(path: Path = DATA_FILE_PATH) -> Dict[str, int]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_CURRENT]

    c_status = _find_header_column(ws, "Status")
    agg: Dict[str, int] = {}

    for r in range(2, ws.max_row + 1):
        s = str(ws.cell(r, c_status).value or "").strip()
        if not s:
            continue
        agg[s] = agg.get(s, 0) + 1

    wb.close()
    return agg
