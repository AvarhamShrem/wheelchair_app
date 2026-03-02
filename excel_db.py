from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import DATA_FILE_PATH, SHEET_CURRENT, SHEET_LISTS, SHEET_LOC, SHEET_TX, SHEET_WC
from models import Transaction


def ensure_workbook_exists(path: Path = DATA_FILE_PATH) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")


def _find_header_column(ws: Worksheet, header_text: str) -> int:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if isinstance(value, str) and value.strip() == header_text:
            return col
    raise KeyError(f"Missing header '{header_text}' in sheet '{ws.title}'")


def list_active_wheelchair_ids(path: Path = DATA_FILE_PATH) -> List[str]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_WC]

    output: List[str] = []
    for row in range(2, ws.max_row + 1):
        wheelchair_id = ws.cell(row, 1).value
        if wheelchair_id is None:
            continue
        wheelchair_id_str = str(wheelchair_id).strip()
        if not wheelchair_id_str:
            continue

        active_flag = ws.cell(row, 9).value
        active_flag_str = str(active_flag).strip().lower() if active_flag is not None else ""
        if active_flag_str in ("", "כן", "true", "1", "y", "yes"):
            output.append(wheelchair_id_str)

    wb.close()
    return output


def list_locations(path: Path = DATA_FILE_PATH) -> List[Dict[str, Any]]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_LOC]

    id_col = _find_header_column(ws, "LocationID")
    name_col = _find_header_column(ws, "LocationName")

    output: List[Dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        location_id = ws.cell(row, id_col).value
        if location_id is None or str(location_id).strip() == "":
            continue
        location_name = ws.cell(row, name_col).value
        output.append({"LocationID": int(location_id), "LocationName": str(location_name or "").strip()})

    wb.close()
    return output


def list_enums(path: Path = DATA_FILE_PATH) -> Dict[str, List[str]]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_LISTS]

    txn_type: List[str] = []
    status: List[str] = []
    condition: List[str] = []

    for row in range(2, ws.max_row + 1):
        txn_value = ws.cell(row, 1).value
        status_value = ws.cell(row, 2).value
        condition_value = ws.cell(row, 3).value

        if txn_value is not None and str(txn_value).strip():
            txn_type.append(str(txn_value).strip())
        if status_value is not None and str(status_value).strip():
            status.append(str(status_value).strip())
        if condition_value is not None and str(condition_value).strip():
            condition.append(str(condition_value).strip())

    wb.close()
    return {"TxnType": txn_type, "StatusAfter": status, "ConditionAfter": condition}


def get_current_location_id(wheelchair_id: str, path: Path = DATA_FILE_PATH) -> Optional[int]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_CURRENT]

    target_id = str(wheelchair_id).strip()
    if not target_id:
        wb.close()
        return None

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if value is None:
            continue
        if str(value).strip() == target_id:
            current = ws.cell(row, 7).value
            wb.close()
            return int(current) if current is not None and str(current).strip() else None

    wb.close()
    return None


def append_transaction(tx: Transaction, path: Path = DATA_FILE_PATH) -> int:
    ensure_workbook_exists(path)
    wb = load_workbook(path)
    ws = wb[SHEET_TX]

    last = ws.max_row
    while last >= 2 and (ws.cell(last, 1).value is None or str(ws.cell(last, 1).value).strip() == ""):
        last -= 1
    new_row = 2 if last < 2 else last + 1

    txn_id = 1 if last < 2 or ws.cell(last, 1).value in (None, "") else int(ws.cell(last, 1).value) + 1

    ws.cell(new_row, 1).value = txn_id
    ws.cell(new_row, 2).value = tx.timestamp
    ws.cell(new_row, 2).number_format = "dd/mm/yyyy hh:mm"
    ws.cell(new_row, 3).value = tx.created_by
    ws.cell(new_row, 4).value = tx.wheelchair_id
    ws.cell(new_row, 5).value = tx.txn_type
    ws.cell(new_row, 6).value = tx.from_location_id
    ws.cell(new_row, 7).value = tx.to_location_id
    ws.cell(new_row, 8).value = tx.status_after
    ws.cell(new_row, 9).value = tx.condition_after or ""
    ws.cell(new_row, 10).value = tx.patient_name or ""
    ws.cell(new_row, 11).value = tx.notes or ""

    ws.cell(new_row, 12).value = f'=IF(AND(A{new_row}<>"",B{new_row}<>""),B{new_row}+A{new_row}/1000000,"")'

    wb.save(path)
    wb.close()
    return txn_id


def count_txn_type_in_range(txn_type: str, start: date, end: date, path: Path = DATA_FILE_PATH) -> int:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_TX]

    timestamp_col = _find_header_column(ws, "Timestamp")
    type_col = _find_header_column(ws, "TxnType")

    count = 0
    for row in range(2, ws.max_row + 1):
        type_value = ws.cell(row, type_col).value
        if str(type_value or "").strip() != txn_type:
            continue

        timestamp_value = ws.cell(row, timestamp_col).value
        if isinstance(timestamp_value, datetime):
            txn_date = timestamp_value.date()
        else:
            try:
                txn_date = datetime.fromisoformat(str(timestamp_value)).date()
            except Exception:
                continue

        if start <= txn_date <= end:
            count += 1

    wb.close()
    return count


def status_aggregation_from_current(path: Path = DATA_FILE_PATH) -> Dict[str, int]:
    ensure_workbook_exists(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_CURRENT]

    status_col = _find_header_column(ws, "Status")
    aggregates: Dict[str, int] = {}

    for row in range(2, ws.max_row + 1):
        status = str(ws.cell(row, status_col).value or "").strip()
        if not status:
            continue
        aggregates[status] = aggregates.get(status, 0) + 1

    wb.close()
    return aggregates
